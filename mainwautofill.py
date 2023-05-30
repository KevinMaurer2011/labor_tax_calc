from openpyxl import load_workbook
import time
import pyautogui
from decimal import *


def autofill1440():
    workbook = load_workbook(filename='rates.xlsx')

    zipint = int(zip_code)
    labor = workbook['labor']
    tax = workbook['tax']
    ziprow = 0
    taxrow = 0

    for i in labor['A']:
        if i.value == zipint:
            ziprow = i.row

    for i in tax['B']:
        if i.value == zipint:
            taxrow = i.row

    try:
        combined_tax = tax.cell(row=taxrow, column=4).value
        tax_percent = (float(combined_tax)) * 100
        tax_rounded = round(tax_percent, 4)
    except:
        print('No tax found sorry')

    # fill in some admin data
    # contact
    pyautogui.moveTo(52, 170, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # estimator
    pyautogui.moveTo(2477, 287, duration=0.5)
    pyautogui.click()
    # jason
    pyautogui.moveTo(2473, 415, duration=0.5)
    pyautogui.click()
    # inspection tab
    pyautogui.moveTo(185, 170, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # open date picker
    pyautogui.moveTo(1477, 259, duration=0.05)
    pyautogui.click()
    # select today
    pyautogui.moveTo(1591, 407, duration=0.5)
    pyautogui.click()
    # site type dropdown
    pyautogui.moveTo(1558, 286, duration=0.5)
    pyautogui.click()
    # virtual
    pyautogui.moveTo(1555, 403, duration=0.5)
    pyautogui.click()
    # settlement tab
    pyautogui.moveTo(493, 171, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # open settlement decision
    pyautogui.moveTo(2168, 473, duration=0.5)
    pyautogui.click()
    # set repairable
    pyautogui.moveTo(2168, 506, duration=0.5)
    pyautogui.click()
    # save file
    pyautogui.moveTo(30, 85, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    # convert to job
    pyautogui.moveTo(296, 91, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    # job number box
    pyautogui.moveTo(1306, 690, duration=0.5)
    pyautogui.click()
    pyautogui.typewrite(file_number)
    pyautogui.typewrite(['enter'])
    time.sleep(3)

    try:
        print('Filling in your rates for you. Standby')
        bodyrate = str(labor.cell(row=ziprow, column=2).value)
        paintrate = str(labor.cell(row=ziprow, column=3).value)
        mechrate = str(labor.cell(row=ziprow, column=4).value)
        framerate = str(labor.cell(row=ziprow, column=5).value)
        pandmrate = str(labor.cell(row=ziprow, column=6).value)

        # click rates
        pyautogui.moveTo(345, 175, duration=0.05)
        pyautogui.click()
        # click labor
        pyautogui.moveTo(40, 200, duration=0.05)
        pyautogui.click()
        # selects labor rates
        pyautogui.moveTo(925, 300, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(bodyrate)
        pyautogui.moveRel(0, 20, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(paintrate)
        pyautogui.moveRel(0, 20, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(mechrate)
        pyautogui.moveRel(0, 20, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(framerate)
        pyautogui.moveRel(0, 160, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(pandmrate)

        labor_taxed_states = (
            'AR', 'CT', 'FL', 'HI', 'IA', 'KS', 'MD', 'MN', 'MS', 'NE', 'NJ', 'NY', 'OH', 'PA', 'SD', 'TX', 'WV', 'WI')

        # if rates are taxed, checks boxes
        if tax.cell(row=taxrow, column=1).value in labor_taxed_states:
            pyautogui.moveTo(997, 298, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 22, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 22, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 22, duration=.05)
            pyautogui.click()
        else:
            print('Labor not Taxed: Skipping')

    except:
        print('No labor rates found')

    try:
        print('Filling in your tax rate for you. Standby')
        # selects taxes tab, and tax rate
        pyautogui.moveTo(215, 200, duration=.05)
        pyautogui.click()
        # for other type tax setup
        pyautogui.moveTo(236, 300, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(str(tax_percent))
        # for traveler type tax setup
        pyautogui.moveTo(400, 320, duration=.05)
        pyautogui.click()
        pyautogui.click()
        pyautogui.typewrite(str(tax_percent))
        pyautogui.typewrite(['enter'])

    except:
        pass

    print(f'Body: {bodyrate}')
    print(f'Paint: {paintrate}')
    print(f'Mech: {mechrate}')
    print(f'Frame: {framerate}')
    print(f'P Supplies: {pandmrate}')
    print(f'Tax Rate: {tax_rounded}')
    print('All done :)')


def autofill1080():
    workbook = load_workbook(filename='rates.xlsx')

    zipint = int(zip_code)
    labor = workbook['labor']
    tax = workbook['tax']
    ziprow = 0
    taxrow = 0

    for i in labor['A']:
        if i.value == zipint:
            ziprow = i.row

    for i in tax['B']:
        if i.value == zipint:
            taxrow = i.row

    try:
        combined_tax = tax.cell(row=taxrow, column=4).value
        tax_percent = (float(combined_tax)) * 100
        tax_rounded = round(tax_percent, 4)
    except:
        print('No tax found sorry')

    # fill in some admin data
    # contact
    pyautogui.moveTo(58, 209, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # estimator
    pyautogui.moveTo(1806, 345, duration=0.5)
    pyautogui.click()
    # jason
    pyautogui.moveTo(1824, 502, duration=0.5)
    pyautogui.click()
    # inspection tab
    pyautogui.moveTo(220, 210, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # open date picker
    pyautogui.moveTo(1190, 316, duration=0.05)
    pyautogui.click()
    # select today
    pyautogui.moveTo(1327, 507, duration=0.5)
    pyautogui.click()
    # site type dropdown
    pyautogui.moveTo(1311, 357, duration=0.5)
    pyautogui.click()
    # virtual
    pyautogui.moveTo(1291, 490, duration=0.5)
    pyautogui.click()
    # settlement tab
    pyautogui.moveTo(576, 208, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # open settlement decision
    pyautogui.moveTo(1697, 578, duration=0.5)
    pyautogui.click()
    # set repairable
    pyautogui.moveTo(1706, 621, duration=0.5)
    pyautogui.click()
    # save file
    pyautogui.moveTo(37, 95, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    # convert to job
    pyautogui.moveTo(333, 98, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    # job number box
    pyautogui.moveTo(1000, 513, duration=0.5)
    pyautogui.click()
    pyautogui.typewrite(file_number)
    pyautogui.typewrite(['enter'])
    time.sleep(3)

    try:
        print('Filling in your rates for you. Standby')
        bodyrate = str(labor.cell(row=ziprow, column=2).value)
        paintrate = str(labor.cell(row=ziprow, column=3).value)
        mechrate = str(labor.cell(row=ziprow, column=4).value)
        framerate = str(labor.cell(row=ziprow, column=5).value)
        pandmrate = str(labor.cell(row=ziprow, column=6).value)

        # click rates
        pyautogui.moveTo(415, 210, duration=0.05)
        pyautogui.click()
        # click labor
        pyautogui.moveTo(40, 240, duration=0.05)
        pyautogui.click()
        # selects labor rates
        pyautogui.moveTo(940, 350, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(bodyrate)
        pyautogui.moveRel(0, 22, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(paintrate)
        pyautogui.moveRel(0, 22, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(mechrate)
        pyautogui.moveRel(0, 22, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(framerate)
        pyautogui.moveRel(0, 170, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(pandmrate)

        labor_taxed_states = (
            'AR', 'CT', 'FL', 'HI', 'IA', 'KS', 'MD', 'MN', 'MS', 'NE', 'NJ', 'NY', 'OH', 'PA', 'SD', 'TX', 'WV', 'WI')

        # if rates are taxed, checks boxes
        if tax.cell(row=taxrow, column=1).value in labor_taxed_states:
            pyautogui.moveTo(997, 345, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 25, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 25, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 25, duration=.05)
            pyautogui.click()
        else:
            print('Labor not Taxed: Skipping')

    except:
        print('No labor rates found')

    try:
        print('Filling in your tax rate for you. Standby')
        # selects taxes tab, and tax rate
        pyautogui.moveTo(245, 235, duration=.05)
        pyautogui.click()
        # for other type tax setup
        pyautogui.moveTo(230, 363, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(str(tax_percent))
        # for traveler type tax setup
        pyautogui.moveTo(425, 375, duration=.05)
        pyautogui.click()
        pyautogui.click()
        pyautogui.typewrite(str(tax_percent))
        pyautogui.typewrite(['enter'])

    except:
        pass

    print(f'Body: {bodyrate}')
    print(f'Paint: {paintrate}')
    print(f'Mech: {mechrate}')
    print(f'Frame: {framerate}')
    print(f'P Supplies: {pandmrate}')
    print(f'Tax Rate: {tax_rounded}')
    print('All done :)')


def get_rates():
    workbook = load_workbook(filename='rates.xlsx')

    zipint = int(zip_code)
    labor = workbook['labor']
    tax = workbook['tax']
    ziprow = 0
    taxrow = 0

    for i in labor['A']:
        if i.value == zipint:
            ziprow = i.row

    for i in tax['B']:
        if i.value == zipint:
            taxrow = i.row

    try:
        combined_tax = tax.cell(row=taxrow, column=4).value
        tax_percent = round(Decimal(combined_tax), 4) * 100

    except:
        print('No tax found sorry')

    try:
        print('Body:  ' + str(labor.cell(row=ziprow, column=2).value))
        print('Paint: ' + str(labor.cell(row=ziprow, column=3).value))
        print('Mech:  ' + str(labor.cell(row=ziprow, column=4).value))
        print('Frame: ' + str(labor.cell(row=ziprow, column=5).value))
        print('P & M: ' + str(labor.cell(row=ziprow, column=6).value))

    except:
        print('No labor rates found')

    try:
        print('Tax: ' + str(tax_percent))

    except:
        pass

    labor_taxed_states = (
        'AR', 'CT', 'FL', 'HI', 'IA', 'KS', 'MD', 'MN', 'MS', 'NE', 'NJ', 'NY', 'OH', 'PA', 'SD', 'TX', 'WV', 'WI')

    if tax.cell(row=taxrow, column=1).value in labor_taxed_states:
        print('Labor Taxed: YES!')
    else:
        print('Labor Taxed: No')
        print('')


program_type = int(input('What feature do you want to use?\n'
                         '1: Autofill - Screen Resolution 1440p\n'
                         '2: Autofill - Screen Resolution 1080p\n'
                         '3: Just Rates - No Autofill\n'
                         'Selection:  '))

while True:
    # autofill 1440p
    if program_type == 1:
        zip_code = input('Zip Code?')
        file_number = input('File Number?')
        autofill1440()
    # autofill 1080p
    if program_type == 2:
        zip_code = input('Zip Code?')
        file_number = input('File Number?')
        #test 
        autofill1080()
    # rates only
    if program_type == 3:
        zip_code = input('Zip Code?')
        #test 
        get_rates()
