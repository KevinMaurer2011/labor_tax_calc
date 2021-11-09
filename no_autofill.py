from openpyxl import load_workbook
import time


def get_rates():
    workbook = load_workbook(filename='rates.xlsx')

    zipint = int(zip_code)
    labor = workbook['labor']
    tax = workbook['tax']
    ziprow = 0
    taxrow = 0

    for i in labor['A']:
        if i.value == zipint:
            ziprow = i.row

    for i in tax['B']:
        if i.value == zipint:
            taxrow = i.row

    try:
        combined_tax = tax.cell(row=taxrow, column=4).value
        tax_percent = (float(combined_tax)) * 100
    except:
        print('No tax found sorry')

    try:
        print('Body:  ' + str(labor.cell(row=ziprow, column=2).value))
        print('Paint: ' + str(labor.cell(row=ziprow, column=3).value))
        print('Mech:  ' + str(labor.cell(row=ziprow, column=4).value))
        print('Frame: ' + str(labor.cell(row=ziprow, column=5).value))
        print('P & M: ' + str(labor.cell(row=ziprow, column=6).value))

    except:
        print('No labor rates found')

    try:
        print('Tax: ' + str(tax_percent))

    except:
        pass

    labor_taxed_states = (
        'AR', 'CT', 'FL', 'HI', 'IA', 'KS', 'MD', 'MN', 'MS', 'NE', 'NJ', 'NY', 'OH', 'PA', 'SD', 'TX', 'WV', 'WI')

    if tax.cell(row=taxrow, column=1).value in labor_taxed_states:
        print('Labor Taxed: YES!')
    else:
        print('Labor Taxed: No')
        print('')


password = input('Whats the password?')
passint = int(password)

if passint == 119911:
    pass
else:
    print('You are a failure')
    time.sleep(2)
    print('My security is top notch')
    time.sleep(2)
    print('Three easy payments of $1,999,999 for full access!')
    time.sleep(2)
    print('I leave now. Bye')
    time.sleep(5)
    quit()

while True:
    zip_code = input('Zip Code?')
    get_rates()
