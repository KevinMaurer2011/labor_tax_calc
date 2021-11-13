from openpyxl import load_workbook
import time
import pyautogui


def get_rates():
    workbook = load_workbook(filename='rates.xlsx')

    zipint = int(zip_code)
    labor = workbook['labor']
    tax = workbook['tax']
    ziprow = 0
    taxrow = 0

    for i in labor['A']:
        if i.value == zipint:
            ziprow = i.row

    for i in tax['B']:
        if i.value == zipint:
            taxrow = i.row

    try:
        combined_tax = tax.cell(row=taxrow, column=4).value
        tax_percent = (float(combined_tax)) * 100
        tax_rounded = round(tax_percent,4)
    except:
        print('No tax found sorry')

    #fill in some admin data
    # contact
    pyautogui.moveTo(58, 209, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # estimator
    pyautogui.moveTo(1806, 345, duration=0.5)
    pyautogui.click()
    # jason
    pyautogui.moveTo(1824, 502, duration=0.5)
    pyautogui.click()
    # inspection tab
    pyautogui.moveTo(220, 210, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # open date picker
    pyautogui.moveTo(1190, 316, duration=0.05)
    pyautogui.click()
    # select today
    pyautogui.moveTo(1327, 507, duration=0.5)
    pyautogui.click()
    # site type dropdown
    pyautogui.moveTo(1311, 357, duration=0.5)
    pyautogui.click()
    # virtual
    pyautogui.moveTo(1335, 430, duration=0.5)
    pyautogui.click()
    # settlement tab
    pyautogui.moveTo(576, 208, duration=0.5)
    pyautogui.click()
    time.sleep(1)
    # open settlement decision
    pyautogui.moveTo(1697, 578, duration=0.5)
    pyautogui.click()
    # set repairable
    pyautogui.moveTo(1706, 621, duration=0.5)
    pyautogui.click()
    # save file
    pyautogui.moveTo(37, 95, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    # convert to job
    pyautogui.moveTo(333, 98, duration=0.5)
    pyautogui.click()
    time.sleep(5)
    # job number box
    pyautogui.moveTo(1000, 513, duration=0.5)
    pyautogui.click()
    pyautogui.typewrite(file_number)
    pyautogui.typewrite(['enter'])
    time.sleep(3)


    try:
        print('Filling in your rates for you. Standby')
        bodyrate = str(labor.cell(row=ziprow, column=2).value)
        paintrate = str(labor.cell(row=ziprow, column=3).value)
        mechrate = str(labor.cell(row=ziprow, column=4).value)
        framerate = str(labor.cell(row=ziprow, column=5).value)
        pandmrate = str(labor.cell(row=ziprow, column=6).value)

        # click rates
        pyautogui.moveTo(415, 210, duration=0.05)
        pyautogui.click()
        # click labor
        pyautogui.moveTo(40, 240, duration=0.05)
        pyautogui.click()
        # selects labor rates
        pyautogui.moveTo(940, 350, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(bodyrate)
        pyautogui.moveRel(0, 22, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(paintrate)
        pyautogui.moveRel(0, 22, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(mechrate)
        pyautogui.moveRel(0, 22, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(framerate)
        pyautogui.moveRel(0, 170, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(pandmrate)


        labor_taxed_states = (
            'AR', 'CT', 'FL', 'HI', 'IA', 'KS', 'MD', 'MN', 'MS', 'NE', 'NJ', 'NY', 'OH', 'PA', 'SD', 'TX', 'WV', 'WI')

        # if rates are taxed, checks boxes
        if tax.cell(row=taxrow, column=1).value in labor_taxed_states:
            pyautogui.moveTo(997, 345, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 25, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 25, duration=.05)
            pyautogui.click()
            pyautogui.moveRel(0, 25, duration=.05)
            pyautogui.click()
        else:
            print('Labor not Taxed: Skipping')

    except:
        print('No labor rates found')

    try:
        print('Filling in your tax rate for you. Standby')
        # selects taxes tab, and tax rate
        pyautogui.moveTo(245, 235, duration=.05)
        pyautogui.click()
        # for other type tax setup
        pyautogui.moveTo(230, 363, duration=.05)
        pyautogui.click()
        pyautogui.typewrite(str(tax_percent))
        # for traveler type tax setup
        pyautogui.moveTo(425, 375, duration=.05)
        pyautogui.click()
        pyautogui.click()
        pyautogui.typewrite(str(tax_percent))
        pyautogui.typewrite(['enter'])

    except:
        pass

    print(f'Body: {bodyrate}')
    print(f'Paint: {paintrate}')
    print(f'Mech: {mechrate}')
    print(f'Frame: {framerate}')
    print(f'P Supplies: {pandmrate}')
    print(f'Tax Rate: {tax_rounded}')
    print('All done :)')



password = input('Whats the password?')
passint = int(password)

if passint == 119911:
    pass
else:
    print('You are a failure')
    time.sleep(2)
    print('My security is top notch')
    time.sleep(2)
    print('Three easy payments of $1,999,999 for full access!')
    time.sleep(2)
    print('I leave now. Bye')
    time.sleep(5)
    quit()

while True:
    zip_code = input('Zip Code?')
    file_number = input('File Number?')
    get_rates()
